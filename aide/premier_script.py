from openpyxl import load_workbook, Workbook
from datetime import datetime
import re

# Charger le classeur source
source_file_name = input("Entrez le nom du fichier Excel source : ")
source_sheet_name = input("Entrez le nom de la feuille de calcul source : ")

source_wb = load_workbook(filename=source_file_name)
source_sheet = source_wb[source_sheet_name]

# Charger ou créer le classeur de destination
destination_file_name = "EB external Jira.xlsx"
try:
    destination_wb = load_workbook(filename=destination_file_name)
except FileNotFoundError:
    destination_wb = Workbook()
    destination_wb.remove(destination_wb.active)  # Supprimer la feuille par défaut

# Créer une nouvelle feuille pour enregistrer les données extraites
if "Innoviz" in destination_wb.sheetnames:
    destination_sheet = destination_wb["Innoviz"]
else:
    destination_sheet = destination_wb.create_sheet("Innoviz")

# Définir les largeurs des colonnes
destination_sheet.column_dimensions['A'].width = 12
destination_sheet.column_dimensions['B'].width = 25
destination_sheet.column_dimensions['C'].width = 100
destination_sheet.column_dimensions['D'].width = 17
destination_sheet.column_dimensions['E'].width = 12
destination_sheet.column_dimensions['F'].width = 12
destination_sheet.column_dimensions['G'].width = 12
destination_sheet.column_dimensions['H'].width = 120
destination_sheet.column_dimensions['I'].width = 25

# En-têtes des colonnes
headers = ['Issue Type', 'Key', 'Summary', 'Resolution', 'Created', 'Updated', 'Resolved', 'Comment (optional)', 'Type']
if destination_sheet.max_row == 1:
    destination_sheet.append(headers)

# Indices des colonnes d'intérêt
indices = {'D': 3, 'B': 1, 'C': 2, 'G': 6, 'K': 10, 'M': 12}

# Table de correspondance pour les mois
month_conversion = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
    "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12"
}

def convert_date(date_str):
    try:
        # Vérifier si la date est au format '27/mar/24'
        day, month_str, year = date_str.split('/')
        month = month_conversion.get(month_str.lower(), month_str)
        return f"{day}/{month}/20{year}"  # Convertir en 'jj/mm/aaaa'
    except ValueError:
        return date_str  # Si la conversion échoue, retourner la date originale

cutoff_date = datetime.strptime("01/06/2022", "%d/%m/%Y")


for row_index, row in enumerate(source_sheet.iter_rows(values_only=True)):
    if row_index in [0, 1, 2, 3]:  # Supprimer les lignes 1, 2, 3 et 4
        continue
    issue = row[indices['D']]
    key = row[indices['B']]
    summary = row[indices['C']]
    resolution = row[indices['G']]
    creation_date = row[indices['K']]
    updated = row[indices['M']]
    resolved = "" 
    comment = ""
    type = ""  

    # Conversion de la date au format approprié
    if isinstance(creation_date, str):
        if re.search('[a-zA-Z]', creation_date):
            formatted_date = convert_date(creation_date.split()[0])
        elif '-' in creation_date:
            year, month, day = creation_date.split('-')
            formatted_date = f"{day}/{month}/{year}"  # Convertir en 'jj/mm/aaaa'
        else:
            formatted_date = creation_date
    elif isinstance(creation_date, datetime):
        formatted_date = creation_date.strftime('%d/%m/%Y')  # Convertir en 'jj/mm/aaaa'
    else:
        formatted_date = creation_date

     # Vérifier si la date de création est après la date limite
    if isinstance(creation_date, str):
        try:
            creation_date_obj = datetime.strptime(formatted_date, "%d/%m/%Y")
        except ValueError:
            continue  # Sauter les lignes avec des dates de création invalides
    elif isinstance(creation_date, datetime):
        creation_date_obj = creation_date
    else:
        continue  # Sauter les lignes avec des dates de création non valides

    if creation_date_obj < cutoff_date:
        continue  # Sauter les lignes avec des dates de création avant la date limite

    # Conversion de la date au format approprié
    if isinstance(updated, str):
        if re.search('[a-zA-Z]', updated):
            formatted_updated = convert_date(updated.split()[0])
        elif '-' in updated:
            year, month, day = updated.split('-')
            formatted_updated = f"{day}/{month}/{year}"  # Convertir en 'jj/mm/aaaa'
        else:
            formatted_updated = updated
    elif isinstance(updated, datetime):
        formatted_updated = updated.strftime('%d/%m/%Y')  # Convertir en 'jj/mm/aaaa'
    else:
        formatted_updated = updated

    # Ajouter les données à la nouvelle feuille de calcul
    row_data = [
        issue,
        key,
        summary,
        resolution,
        formatted_date,
        formatted_updated,
        resolved,  
        comment,
        type
    ]
    destination_sheet.append(row_data)

    # Ajouter la formule "IF" dans la colonne "Resolved"
    formula = f'=IF(D{destination_sheet.max_row}="Fixed", F{destination_sheet.max_row}, "NO")'
    destination_sheet.cell(row=destination_sheet.max_row, column=7).value = formula 

# Enregistrer le classeur modifié avec les données extraites
destination_wb.save(destination_file_name)

print(f"Les données ont été extraites et enregistrées dans '{destination_file_name}'.")
from openpyxl import load_workbook, Workbook
from openpyxl.utils import FORMULAE
from datetime import datetime
import re
import pandas as pd

# Charger le classeur source
source_file_name = input("Entrez le nom du fichier Excel source : ")
source_sheet_name = input("Entrez le nom de la feuille de calcul source : ")

source_wb = load_workbook(filename=source_file_name)
source_sheet = source_wb[source_sheet_name]

# Charger ou créer le classeur de destination
destination_file_name = "JIRA_summary.xlsx"
try:
    destination_wb = load_workbook(filename=destination_file_name)
except FileNotFoundError:
    destination_wb = Workbook()
    destination_wb.remove(destination_wb.active)  # Supprimer la feuille par défaut

# Créer une nouvelle feuille pour enregistrer les données extraites
if "general_report" in destination_wb.sheetnames:
    destination_sheet = destination_wb["general_report"]
else:
    destination_sheet = destination_wb.create_sheet("general_report")

destination_sheet.column_dimensions['A'].width = 22
destination_sheet.column_dimensions['B'].width = 90
destination_sheet.column_dimensions['C'].width = 12
destination_sheet.column_dimensions['D'].width = 25
destination_sheet.column_dimensions['E'].width = 12
destination_sheet.column_dimensions['F'].width = 12

# En-têtes des colonnes
headers = ['Key', 'Summary', 'Status', 'Assignee', 'Created', 'Updated']
if destination_sheet.max_row == 1:
    destination_sheet.append(headers)

# Indices des colonnes d'intérêt
indices = {'B': 1, 'C': 2, 'E': 4, 'H': 7, 'K': 10, 'M': 12}

# Table de correspondance pour les mois
month_conversion = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
    "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12"
}

def convert_date(date_str):
    try:
        # Vérifier si la date est au format '27/mar/24'
        day, month_str, year = date_str.split('/')
        month = month_conversion.get(month_str.lower(), month_str)
        return f"{day}/{month}/20{year}"  # Convertir en 'jj/mm/aaaa'
    except ValueError:
        return date_str  # Si la conversion échoue, retourner la date originale

cutoff_date = datetime.strptime("01/06/2022", "%d/%m/%Y")

# Récupérer les informations de la feuille de destination
destination_data = {}
for row in destination_sheet.iter_rows(min_row=2, values_only=False):  # values_only=False pour accéder aux hyperliens
    key = row[0].value
    status = row[2].value
    if key:
        destination_data[key] = row[0].row

# Vérifier et supprimer les entrées passées à "Resolved" ou "Closed"
for row_index, row in enumerate(source_sheet.iter_rows(values_only=False)):
    if row_index in [0, 1, 2, 3]:  # Ignorer les premières lignes
        continue
    key = row[indices['B']].value
    status = row[indices['E']].value

    if key in destination_data and status in ['Resolved', 'Closed']:
        destination_sheet.delete_rows(destination_data[key], 1)
        del destination_data[key]


for row_index, row in enumerate(source_sheet.iter_rows(values_only=False)):  # Changer values_only à False pour accéder aux hyperliens
    if row_index in [0, 1, 2, 3]:  # Supprimer les lignes 1, 2, 3 et 4
        continue
    key = row[indices['B']].value
    summary = row[indices['C']].value
    status = row[indices['E']].value
    assignee = row[indices['H']].value
    creation_date = row[indices['K']].value
    updated = row[indices['M']].value

    # Ignorer les lignes où le status est "Resolved" ou "Closed"
    if status in ['Resolved', 'Closed']:
        continue

    # Ignorer les lignes où la clé existe déjà
    if key in destination_data:
        continue

    # Conversion de la date au format approprié
    if isinstance(creation_date, str):
        if re.search('[a-zA-Z]', creation_date):
            formatted_date = convert_date(creation_date.split()[0])
        elif '-' in creation_date:
            year, month, day = creation_date.split('-')
            formatted_date = f"{day}/{month}/{year}"  # Convertir en 'jj/mm/aaaa'
        else:
            formatted_date = creation_date
    elif isinstance(creation_date, datetime):
        formatted_date = creation_date.strftime('%d/%m/%Y')  # Convertir en 'jj/mm/aaaa'
    else:
        formatted_date = creation_date

    # Conversion de la date au format approprié
    if isinstance(updated, str):
        if re.search('[a-zA-Z]', updated):
            formatted_updated = convert_date(updated.split()[0])
        elif '-' in updated:
            year, month, day = updated.split('-')
            formatted_updated = f"{day}/{month}/{year}"  # Convertir en 'jj/mm/aaaa'
        else:
            formatted_updated = updated
    elif isinstance(updated, datetime):
        formatted_updated = updated.strftime('%d/%m/%Y')  # Convertir en 'jj/mm/aaaa'
    else:
        formatted_updated = updated

    # Ajouter les données à la nouvelle feuille de calcul
    row_data = [
        key,
        summary,
        status,
        assignee,
        formatted_date,
        formatted_updated,
    ]

    destination_sheet.append(row_data)

    # Ajouter les hyperliens si présents
    if row[indices['B']].hyperlink:
        destination_sheet.cell(row=destination_sheet.max_row, column=1).hyperlink = row[indices['B']].hyperlink

    # Ajouter la clé à l'ensemble des clés existantes
    destination_data[key] = destination_sheet.max_row

# Enregistrer le classeur modifié avec les données extraites
destination_wb.save(destination_file_name)

print(f"Les données ont été extraites et enregistrées dans '{destination_file_name}'.")